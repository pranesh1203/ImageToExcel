import os,re
from msrest.authentication import CognitiveServicesCredentials
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from azure.cognitiveservices.vision.computervision.models import OperationStatusCodes
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinterdnd2 import *
class ImgtoExl:
    def __init__(self,s):
        s1=s.replace('{','')
        s2=s1.replace('}','')
        self.s=s2
        API_KEY="545a2ccf0b184086acdc5322e0efe843"
        ENDPOINT="https://pranesh45.cognitiveservices.azure.com/"
        computervision_client=ComputerVisionClient(ENDPOINT,CognitiveServicesCredentials(API_KEY))
        read_image = open(self.s, "rb")
        response = computervision_client.read_in_stream(read_image,raw=True)
        operationlocation=response.headers['Operation-Location']
        operation_id=operationlocation.split('/')[-1]
        while True:
            read_result = computervision_client.get_read_result(operation_id)
            if read_result.status.lower () not in ['notstarted', 'running']:
                break
        l=[]
        if read_result.status == OperationStatusCodes.succeeded:
            read_results=read_result.analyze_result.read_results
            for text_result in read_results:
                for line in text_result.lines:
                    l.append(line.text)
        r=req(l)
        l4=[]
        for i in r:
            l4.append(req2(i))
        self.l=l4
        print(l4)
    def new(self):
        wb=Workbook()
        ws=wb.active
        ws['A1'].value='Lab Marks Evaluation'
        ws.merge_cells("A1:T1")
        ws['A2'].value='Roll-Number'
        ws['B2'].value='Marks'
        ws.merge_cells("B2:R2")
        for col in range(2,18,2):
            c=get_column_letter(col)
            ws[c+'4'].value="LAB"
        for col in range(3,19,2):
            c=get_column_letter(col)
            ws[c+'4'].value="VIVA"
        i=1
        for col in range(2,18,2):
            c=get_column_letter(col)
            ws[c+'3'].value=f"Week-{i}"
            i=i+1
        ws.merge_cells("A2:A4")
        for i in self.l:
            ws.append(i)
        x=os.path.basename(self.s)
        x=x.split('.')
        wb.save(f'{x[0]}.xlsx')
    def upd(self,a):
        a1=a.replace('{','')
        a2=a1.replace('}','')
        self.a=a2
        wb=load_workbook(self.a)
        ws=wb.active
        for i in self.l:
            ws.append(i)
        x=os.path.basename(self.a)
        x=x.split('.')
        wb.save(f'{x[0]}.xlsx')
    def new_win(self):
        main=Tk()
        main.title('Image to Excel')
        main.geometry('400x300')
        main.config(bg='white')
        Label(main,text='Edit and Confirm').pack()

def req(l):
    s1=''
    l1=[]
    for i in l:
        s1=re.sub("[^0-9,-]","",i)
        l1.append(s1)
    l1=l1[25:]
    l2=[]
    i=0
    for i in range(len(l1)):
        if(len(l1[i])>15 and re.search('^1602',l1[i])):
            l=[]
            s1=l1[i][0:15]
            s2=l1[i][15:]
            l2.append([s1,s2])
        elif(len(l1[i])>5 and re.search('^1602',l1[i])):
                if(len(l1[i+1])>15):
                    l2.append([l1[i],l1[i+1]])
        else:
            continue
    return l2
def req2(l):
    s=l[1].split(',')
    s.pop(0)
    l.pop()
    for i in s:
        if i=='':
            s.remove('')
        else:
            l.append(int(i))
    return l
def addto_listbox(event):
    lb1.insert("end", event.data)
def addto_listbox1(event):
    lb2.insert("end", event.data)
def showSelected():
    x1= []
    for i in range(lb1.size()):
        op=lb1.get(i)
        x1.append(op)
    obj1=ImgtoExl(x1[0])
    obj1.new()
def showSelected1():
    x2=[]
    for i in range(lb2.size()):
        op = lb2.get(i)
        x2.append(op)
    obj2=ImgtoExl(x2[1])
    obj2.upd(x2[0])
ws = TkinterDnD.Tk()
ws.title('Image to Excel')
ws.geometry('400x300')
ws.config(bg='white')
l1=Label(ws,text="Drag and drop Image")
l1.pack()
frame = Frame(ws)
frame.pack()
lb1 = Listbox(frame, width=30,height=3,selectmode=MULTIPLE)
lb1.pack(fill=X, side=LEFT)
lb1.drop_target_register(DND_FILES)
lb1.dnd_bind('<<Drop>>', addto_listbox)
Button(ws, text="Upload", command=showSelected).pack()
l2=Label(ws,text="Drag and drop Excel and Image")
l2.pack()
f2 = Frame(ws)
f2.pack()
lb2 = Listbox(f2, width=30,height=3,selectmode=MULTIPLE)
lb2.pack(fill=X, side=LEFT)
lb2.drop_target_register(DND_FILES)
lb2.dnd_bind('<<Drop>>', addto_listbox1)
Button(ws, text="Update", command=showSelected1).pack()
ws.mainloop()