from PyQt5 import QtCore, QtGui, QtWidgets
import re,os
from PyQt5.QtWidgets import QMainWindow, QListWidget, QListWidgetItem,QLabel
import pandas as pd
from PyQt5.QtCore import Qt
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from azure.cognitiveservices.vision.computervision.models import OperationStatusCodes
from msrest.authentication import CognitiveServicesCredentials
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QWidget, QTableWidget, QTableWidgetItem, QHeaderView,QPushButton, QVBoxLayout
lan=[]
lan2=[]
imgadd=''
exladd=''
class TableWidget(QTableWidget):   #TableWidget
    def __init__(self, df):
        super().__init__()
        self.df = df
        self.setStyleSheet('font-size: 15px;')

        # set table dimension
        nRows, nColumns = self.df.shape
        self.setColumnCount(nColumns)
        self.setRowCount(nRows)
        self.setHorizontalHeaderLabels(('Name', 'Scores'))
        self.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # data insertion
        for i in range(self.rowCount()):
            for j in range(self.columnCount()):
                self.setItem(i, j, QTableWidgetItem(str(self.df.iloc[i, j])))

        self.cellChanged[int, int].connect(self.updateDF)

    def updateDF(self, row, column):
        text = self.item(row, column).text()
        self.df.iloc[row, column] = text # updates the new input into the data frame


def dic(l):
    l1 = []
    l2 = []
    for i in l:
        while (len(i) < 17):
            i.append('-') # appends hyphen in case of empty blocks
    for i in l:
        s3 = ''
        l1.append(i[0])
        for j in range(1, len(i)):
            s3 += f"{i[j]},"
        l2.append(s3)
    return [l1,l2] # list to view data on table widget
class DFEditor(QWidget):
    def __init__(self):
        super().__init__()
        self.l = lan2
        self.lst2 = dic(self.l)
        data = {
            'Roll-Number': self.lst2[0], # list of roll numbers
            'Score': self.lst2[1] # list of marks
        }
        df = pd.DataFrame(data)
        self.resize(800,600)

        man = QVBoxLayout()

        self.table = TableWidget(df) # creating TableWidget with data frame

        man.addWidget(self.table)

        button_print = QPushButton('Update') # updates the corrected data from table widget
        button_print.setStyleSheet('font-size: 30px')
        button_print.clicked.connect(self.print_DF_Values)
        man.addWidget(button_print)
        self.setLayout(man)

    def print_DF_Values(self):
        global lan
        lan = self.df.values.tolist()
        l2=req3(lan)

def f2():  # instance for Table widget window
    w = DFEditor()
    w.show()
def req3(l): # list for table widget
    l1=[]
    for i in l:
        l2=[]
        l2.append(i[0])
        s=i[1].split(',')
        for j in s:
            if j=='':
                s.remove('')
            else:
                l2.append(int(j))
        l1.append(l2)
    return l1 #lst=[['rol',m1,m2,m3,m4,....],.....]
class ImgtoExl(QMainWindow):
    def __init__(self,s):
        super().__init__()
        s1=s.replace('{','')     # eliminating the curly braces
        s2=s1.replace('}','')
        self.s=s2       # addreass of Image
        API_KEY=""# Enter   API KEY here
        ENDPOINT="" # Enter ENDPOINT here
        computervision_client=ComputerVisionClient(ENDPOINT,CognitiveServicesCredentials(API_KEY))
        read_image = open(self.s, "rb")
        response = computervision_client.read_in_stream(read_image,raw=True)  # api call
        operationlocation=response.headers['Operation-Location']
        operation_id=operationlocation.split('/')[-1]
        while True:   # waiting for api results
            read_result = computervision_client.get_read_result(operation_id)
            if read_result.status.lower () not in ['notstarted', 'running']:
                break
        l=[]
        if read_result.status == OperationStatusCodes.succeeded:  # appending the extracted data into a list
            read_results=read_result.analyze_result.read_results
            for text_result in read_results:
                for line in text_result.lines:
                    print(line)
                    l.append(line.text)
        r=req(l)   # get a list consisting of list of roll num and marks
        l4=[]
        for i in r:
            l4.append(req2(i)) # convert marks from str to int and append into new list
        global lan2
        lan2 = l4
        f2()  # call table widget window
        self.l=l4
    def new(self): # method to add data into new excdel sheet
        wb=Workbook() # Opening a new excel sheet
        ws=wb.active # function to access the excel functions and cells
        ws['A1'].value='Lab Marks Evaluation' # column names
        ws.merge_cells("A1:T1")
        ws['A2'].value='Roll-Number'
        ws['B2'].value='Marks'
        ws.merge_cells("B2:R2")
        for col in range(2,18,2):
            c=get_column_letter(col)
            ws[c+'4'].value="LAB"
        for col in range(3,19,2):
            c=get_column_letter(col)
            ws[c+'4'].value="VIVA"
        i=1
        for col in range(2,18,2):
            c=get_column_letter(col)
            ws[c+'3'].value=f"Week-{i}"
            i=i+1
        ws.merge_cells("A2:A4")
        for i in self.l:    # appends each list inside the list to the excel sheet as a row
            ws.append(i)
        x=os.path.basename(self.s) # gets the name of image
        x=x.split('.')
        wb.save(f'{x[0]}.xlsx') # saving the excel sheet on the name of uploaded image
    def upd(self,a): # method to update data into existing excel sheet
        a1=a.replace('{','')
        a2=a1.replace('}','')
        self.a=a2 # address of existing excel sheet
        wb=load_workbook(self.a) # loading existing excel sheet
        ws=wb.active
        for i in self.l:
            ws.append(i)
        x=os.path.basename(self.a)
        x=x.split('.')
        wb.save(f'{x[0]}.xlsx') # saving the newly added data to excel sheet

def req(l):
    s1=''
    l1=[]
    for i in l:
        s1=re.sub("[^0-9,-]","",i) # remove all other elements except numbers and commas
        l1.append(s1)
    l1=l1[25:] # removing the unwanted data
    l2=[]
    for i in range(len(l1)):
        if(len(l1[i])>15 and re.search('^1602',l1[i])):
            l=[]
            s1=l1[i][0:15]
            s2=l1[i][15:]
            l2.append([s1,s2])
        elif(len(l1[i])>5 and re.search('^1602',l1[i])):
                if(len(l1[i+1])>15):
                    l2.append([l1[i],l1[i+1]])
        else:
            continue
    return l2 # lst=[['roll','m1,m2,m3,...'],....]
def req2(l):
    s=l[1].split(',') # list of string of numbers
    s.pop(0) # removes the comma at the start
    l.pop() # remove the string of marks in original list
    for i in s:
        if i=='':
            s.remove('') # removes blank index
        else:
            l.append(int(i)) # appends integer into original list
    return l # lst=[['rol',m1,m2,m3,m4,....],.....]
def f():        # Creating an instance for Uploading Window and display
    demo = AppDemo()
    demo.show()
class ListBoxWidget(QListWidget):        # List widget to drag and drop files
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.resize(300,300)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()

            links = []
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    links.append(str(url.toLocalFile()))
                else:
                    links.append(str(url.toString()))
            self.addItems(links)
        else:
            event.ignore()

class AppDemo(QMainWindow):          # Window to Drag n Drop files
    def __init__(self):
        super().__init__()
        self.resize(800, 600)
        self.label=QLabel('Drag and Drop files here',self)
        self.label.setFixedWidth(200)
        self.label.move(325,50)
        self.listbox_view = ListBoxWidget(self)
        self.listbox_view.setGeometry(125,100,500,300)
        self.btn = QPushButton('Upload', self)
        self.btn.setGeometry(300,400,200,50)
        self.btn.clicked.connect(lambda: (self.getSelectedItem())) # Upload files button

    def getSelectedItem(self):
        self.l=[]
        for i in range(self.listbox_view.count()):
            self.item = QListWidgetItem(self.listbox_view.item(i))
            self.l.append(self.item.text())
        if(len(self.l)==1):   # Passing address of image to class for API call
            obj1=ImgtoExl(self.l[0])
            obj1.new()        # Method to create a new excel sheet with data
        else:                 # Passing address of image to class for API call
            obj2 = ImgtoExl(self.l[1])
            obj2.upd(self.l[0])   # Passing address of excel file to load existing sheet

class Ui_MainWindow(object):             #Start Window
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        MainWindow.setFont(font)
        MainWindow.setAutoFillBackground(False)
        MainWindow.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0.04, x2:0.706, y2:0.21, stop:0 rgba(25, 139, 48, 255), stop:1 rgba(0, 0, 0, 255));")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(200, 70, 391, 141))
        self.label.setStyleSheet("border-style:outset;\n"
"border-width:2px;\n"
"border-radius:10px;")
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("D:\\C-Vscode\\Mini Project\\images/Capture3.PNG"))
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(340, 310, 131, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color:rgb(0, 145, 5);\n"
"color: rgb(255, 255, 255);\n"
"border-style:outset;\n"
"border-width:2px;\n"
"border-radius:10px;\n"
"border-color: rgb(255, 255, 255);")
        icon = QtGui.QIcon.fromTheme("Upload")
        self.pushButton.setIcon(icon)
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(340, 380, 131, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("background-color:rgb(0, 145, 5);\n"
"color: rgb(255, 255, 255);\n"
"border-style:outset;\n"
"border-width:2px;\n"
"border-radius:10px;\n"
"background-color: rgb(0, 0, 0);")
        self.pushButton_2.setObjectName("pushButton_2")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.pushButton.clicked.connect(f)      # Open Uploading window when clicked
        self.pushButton_2.clicked.connect(f)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.pushButton.setText(_translate("MainWindow", "Upload"))
        self.pushButton_2.setText(_translate("MainWindow", "Update"))


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    app.exec_()
